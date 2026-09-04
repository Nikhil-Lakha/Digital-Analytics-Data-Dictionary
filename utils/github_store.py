import base64
from io import BytesIO
from pathlib import Path

import requests
from openpyxl import load_workbook

REPO = "Nikhil-Lakha/Digital-Analytics-Data-Dictionary"
FILE_PATH = "data/analytics_data_dictionary.xlsx"
BRANCH = "main"
API_URL = f"https://api.github.com/repos/{REPO}/contents/{FILE_PATH}"
RAW_URL = f"https://raw.githubusercontent.com/{REPO}/{BRANCH}/{FILE_PATH}"
LOCAL_XLSX_PATH = Path(__file__).resolve().parents[1] / FILE_PATH


def _headers(token: str | None = None) -> dict:
    headers = {
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "digital-analytics-data-dictionary",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def fetch_workbook_bytes(token: str | None = None) -> bytes:
    """Fetch the latest workbook from GitHub."""
    if token:
        response = requests.get(API_URL, headers=_headers(token), params={"ref": BRANCH}, timeout=20)
        response.raise_for_status()
        payload = response.json()
        return base64.b64decode(payload["content"])

    response = requests.get(RAW_URL, headers=_headers(), timeout=20)
    response.raise_for_status()
    return response.content


def _get_file_metadata(token: str) -> tuple[str, bytes]:
    response = requests.get(API_URL, headers=_headers(token), params={"ref": BRANCH}, timeout=20)
    response.raise_for_status()
    payload = response.json()
    return payload["sha"], base64.b64decode(payload["content"])


def _commit_workbook(token: str, workbook_bytes: bytes, message: str) -> None:
    sha, _ = _get_file_metadata(token)
    payload = {
        "message": message,
        "content": base64.b64encode(workbook_bytes).decode("utf-8"),
        "sha": sha,
        "branch": BRANCH,
    }
    response = requests.put(API_URL, headers=_headers(token), json=payload, timeout=30)
    response.raise_for_status()


def _load_for_write(token: str | None):
    """Use GitHub when a token is supplied; otherwise use the local workbook for localhost testing."""
    if token:
        _, workbook_bytes = _get_file_metadata(token)
        return load_workbook(BytesIO(workbook_bytes))

    if not LOCAL_XLSX_PATH.exists():
        raise FileNotFoundError(f"Local workbook was not found at {LOCAL_XLSX_PATH}")
    return load_workbook(LOCAL_XLSX_PATH)


def _save_after_write(token: str | None, wb, message: str) -> None:
    if token:
        output = BytesIO()
        wb.save(output)
        _commit_workbook(token, output.getvalue(), message)
    else:
        LOCAL_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb.save(LOCAL_XLSX_PATH)


def _headers_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): idx
        for idx, cell in enumerate(ws[1], start=1)
        if cell.value is not None
    }


def _find_variable_row(ws, variable_name: str) -> tuple[int, dict[str, int]]:
    headers = _headers_map(ws)
    variable_col = headers.get("Variable Name")
    if not variable_col:
        raise ValueError("Variable Name column is missing from the workbook.")

    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=variable_col).value
        if str(value).strip() == str(variable_name).strip():
            return row_idx, headers

    raise ValueError(f"Variable '{variable_name}' was not found in the workbook.")


def create_variable(token: str | None, values: dict) -> None:
    wb = _load_for_write(token)
    ws = wb["Variables"]
    headers = _headers_map(ws)

    variable_name = str(values.get("Variable Name", "")).strip()
    if not variable_name:
        raise ValueError("Variable Name is required.")

    variable_col = headers.get("Variable Name")
    for row_idx in range(2, ws.max_row + 1):
        existing = ws.cell(row=row_idx, column=variable_col).value
        if str(existing).strip().lower() == variable_name.lower():
            raise ValueError(f"Variable '{variable_name}' already exists.")

    new_row = ws.max_row + 1
    for field, col_idx in headers.items():
        ws.cell(row=new_row, column=col_idx).value = values.get(field, "")

    _save_after_write(token, wb, f"Add analytics variable: {variable_name}")


def update_variable(token: str | None, original_variable_name: str, values: dict) -> None:
    wb = _load_for_write(token)
    ws = wb["Variables"]
    row_idx, headers = _find_variable_row(ws, original_variable_name)

    new_variable_name = str(values.get("Variable Name", original_variable_name)).strip()
    if new_variable_name.lower() != str(original_variable_name).strip().lower():
        variable_col = headers.get("Variable Name")
        for check_row in range(2, ws.max_row + 1):
            if check_row == row_idx:
                continue
            existing = ws.cell(row=check_row, column=variable_col).value
            if str(existing).strip().lower() == new_variable_name.lower():
                raise ValueError(f"Variable '{new_variable_name}' already exists.")

    for field, value in values.items():
        col_idx = headers.get(field)
        if col_idx:
            ws.cell(row=row_idx, column=col_idx).value = value

    _save_after_write(token, wb, f"Update analytics variable: {original_variable_name}")


def delete_variable(token: str | None, variable_name: str) -> None:
    wb = _load_for_write(token)
    ws = wb["Variables"]
    row_idx, _ = _find_variable_row(ws, variable_name)
    ws.delete_rows(row_idx, 1)

    _save_after_write(token, wb, f"Delete analytics variable: {variable_name}")
